import openpyxl

#File--->Workbook--->Sheets--->Rows--->Columns

File="C:\SeleniumPractice\data.xlsx"
Workbook=openpyxl.load_workbook(File)
Sheet=Workbook["Sheet1"]

Rows=Sheet.max_row
Cols=Sheet.max_column
for r in range(1,Rows+1):
    for c in range(1,Cols+1):
        print(Sheet.cell(r,c).value,end="      ")
    print()
