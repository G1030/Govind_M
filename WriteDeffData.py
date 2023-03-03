import openpyxl

file="C:\\SeleniumPractice\\diffrentdata.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook.active

sheet.cell(1,1).value=123
sheet.cell(1,2).value="Govind"
sheet.cell(1,3).value="Manager"

sheet.cell(2,1).value=345
sheet.cell(2,2).value="Raj"
sheet.cell(2,3).value="software Tester"

sheet.cell(3,1).value=567
sheet.cell(3,2).value="Vijay"
sheet.cell(3,3).value="Developer"

workbook.save(file)
print("Writing Different Data Success!!!")


